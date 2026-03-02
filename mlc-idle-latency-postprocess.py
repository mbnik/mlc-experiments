#!/usr/bin/env python3
"""
MLC Idle Latency Post-Processing Script

Reads a summary.csv from an MLC idle latency sweep results folder and generates
an Excel report with per-buffer-size iteration data and aggregated statistics.

Prerequisites:
    conda activate mlc_3p12

Usage:
    python mlc-idle-latency-postprocess.py --results-dir results/20260301/idle_latency/node0
"""

import argparse
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


METRIC_COLUMNS = [
    ("clocks", "Clocks"),
    ("latency_ns", "Latency (ns)"),
]

STAT_NAMES = ["mean", "std", "min", "max"]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Post-process MLC idle latency sweep results into an Excel report"
    )
    parser.add_argument(
        "--results-dir", required=True,
        help="Path to results folder containing summary.csv"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output Excel file path (default: <results-dir>/summary.xlsx)"
    )
    return parser.parse_args()


def load_data(results_dir):
    """Load and validate summary.csv."""
    csv_path = os.path.join(results_dir, "summary.csv")
    if not os.path.isfile(csv_path):
        print(f"ERROR: summary.csv not found in {results_dir}")
        sys.exit(1)

    df = pd.read_csv(csv_path)
    required = {"buffer_size", "core", "iteration", "clocks", "latency_ns"}
    missing = required - set(df.columns)
    if missing:
        print(f"ERROR: Missing columns in CSV: {missing}")
        sys.exit(1)

    return df


def write_section_header(ws, row, text, num_cols):
    """Write a bold section header merged across columns."""
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def write_table_header(ws, row):
    """Write column headers for a data block."""
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin"))

    headers = ["Iter"] + [display for _, display in METRIC_COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    return row + 1


def write_data_rows(ws, row, config_df):
    """Write iteration data rows."""
    num_fmt = "#,##0.00"
    for _, data_row in config_df.iterrows():
        ws.cell(row=row, column=1, value=int(data_row["iteration"])).alignment = Alignment(horizontal="center")
        for col_idx, (csv_col, _) in enumerate(METRIC_COLUMNS, start=2):
            cell = ws.cell(row=row, column=col_idx, value=data_row[csv_col])
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal="center")
        row += 1
    return row


def write_stat_rows(ws, row, config_df):
    """Write mean/std/min/max rows."""
    num_fmt = "#,##0.00"
    stat_font = Font(bold=True, size=10)
    stat_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    metric_cols = [col for col, _ in METRIC_COLUMNS]
    stats = config_df[metric_cols].agg(STAT_NAMES)

    for stat_name in STAT_NAMES:
        cell = ws.cell(row=row, column=1, value=stat_name)
        cell.font = stat_font
        cell.fill = stat_fill
        cell.alignment = Alignment(horizontal="center")
        for col_idx, csv_col in enumerate(metric_cols, start=2):
            cell = ws.cell(row=row, column=col_idx, value=stats.loc[stat_name, csv_col])
            cell.number_format = num_fmt
            cell.font = stat_font
            cell.fill = stat_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1
    return row


def auto_width(ws):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)


def build_report(df, output_path):
    """Build the Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MLC Idle Latency Results"

    buffer_sizes = list(dict.fromkeys(df["buffer_size"]))
    num_cols = 1 + len(METRIC_COLUMNS)
    n_iters = df["iteration"].nunique()

    row = 1

    # --- Info header ---
    device_id = df["device_id"].iloc[0] if "device_id" in df.columns else "?"
    core = df["core"].iloc[0] if "core" in df.columns else "?"
    info_font = Font(italic=True, size=10, color="555555")
    ws.cell(row=row, column=1,
            value=f"Device: node{device_id}  |  Core: {core}  |  Buffer sizes: {len(buffer_sizes)}  |  Iterations: {n_iters}")
    ws.cell(row=row, column=1).font = info_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 2

    # ===================== Raw Data Section =====================
    row = write_section_header(ws, row, "Raw Data", num_cols)
    row += 1

    title_font = Font(bold=True, size=10, color="2F5496")

    for buf in buffer_sizes:
        config_df = df[df["buffer_size"] == buf].sort_values("iteration")

        cell = ws.cell(row=row, column=1, value=f"Buffer: {buf}")
        cell.font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        row += 1

        row = write_table_header(ws, row)
        row = write_data_rows(ws, row, config_df)
        row += 1

    # ===================== Summary Section =====================
    row += 1
    row = write_section_header(ws, row, "Summary (Aggregated Statistics)", num_cols)
    row += 1

    for buf in buffer_sizes:
        config_df = df[df["buffer_size"] == buf]

        cell = ws.cell(row=row, column=1, value=f"Buffer: {buf}")
        cell.font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        row += 1

        row = write_table_header(ws, row)
        row = write_stat_rows(ws, row, config_df)
        row += 1

    # ===================== Latency Table =====================
    row += 1
    row = write_section_header(ws, row, "Latency Summary Table (mean ns per buffer size)", num_cols)
    row += 1

    # Header
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin"))

    for col_idx, header in enumerate(["Buffer Size", "Mean Latency (ns)"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Data
    means = df.groupby("buffer_size", sort=False)["latency_ns"].mean()
    for buf in buffer_sizes:
        ws.cell(row=row, column=1, value=buf).font = Font(bold=True, size=10)
        cell = ws.cell(row=row, column=2, value=means[buf])
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="center")
        row += 1

    auto_width(ws)
    wb.save(output_path)
    print(f"Report saved to {output_path}")


def main():
    args = parse_args()

    if args.output is None:
        args.output = os.path.join(args.results_dir, "summary.xlsx")

    df = load_data(args.results_dir)
    print(f"Loaded {len(df)} rows from {args.results_dir}/summary.csv")
    print(f"  Buffer sizes: {df['buffer_size'].nunique()}")
    print(f"  Iterations: {df['iteration'].nunique()}")
    print()

    build_report(df, args.output)


if __name__ == "__main__":
    main()
